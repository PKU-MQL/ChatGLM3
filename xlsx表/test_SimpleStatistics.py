import pandas as pd
import numpy as np  
from openpyxl import load_workbook

'''
背景知识：/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx文件中
共有4个独立的sheet，分别为“民用汽车拥有量汇总”、“民用载货”、“民用载客”、“民用其他”。
“民用汽车拥有量汇总”、“民用载货”、“民用载客”的列名都从第5行开始，
列名有省份、2002年、2003年……一直到2020年等项，不存在汇总各个省份的全国项。
单位为万辆，根据下列查询要求，输出实现代码。
'''



def testss01():#计算出每个省份2020年相较于2010年的民用汽车拥有量增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 判断数据中是否存在'2010年'这一列  
    if '2010年' in df.columns:  
        # 计算每个省份2020年相较于2010年的民用汽车拥有量增长率  
        df['增长率'] = (df['2020年'] - df['2010年']) / df['2010年'] * 100  
    else:  
        print("数据中不存在'2010年'这一列")  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的增长率数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['增长率'])  # 将增长率写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def testss02():#计算出每个省份2013年相较于2002年的民用汽车拥有量增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 判断数据中是否存在'2002年'这一列  
    if '2002年' in df.columns:  
        # 计算每个省份2013年相较于2002年的民用汽车拥有量增长率  
        df['增长率'] = (df['2013年'] - df['2002年']) / df['2002年'] * 100  
    else:  
        print("数据中不存在'2002年'这一列")  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的增长率数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['增长率'])  # 将增长率写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def testss03():#将各省份按照2010年的民用汽车拥有量降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 按照2010年的民用汽车拥有量进行降序排序  
    df.sort_values('2010年', inplace=True, ascending=False)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将排序后的数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（覆盖原有数据）  
    for index, row in df.iterrows():  
        for col_num, value in enumerate(row.values, start=1):  
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)

def testss04():#将各省份按照2010年的民用汽车拥有量升序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 对各省份按照2010年的民用汽车拥有量进行升序排序  
    df.sort_values('2010年', inplace=True)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将排序后的数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（覆盖原有数据）  
    for index, row in df.iterrows():  
        for col_num, value in enumerate(row.values, start=1):  
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)

def testss05():#计算出这20年中每年各个省份的民用载货汽车拥有量的平均值，结果存储在第40行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用载货',engine='openpyxl',header=4)
    # 计算每年各个省份的民用载货汽车拥有量的平均值  
    average_df = df.iloc[:, 1:].mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用载货'这张表  
    ws = book["民用载货"]  
  
    # 将结果写入原Excel文件的'民用载货'sheet中的第40行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=40, column=col_num, value=value)  # 写入平均值到第40行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

'''
背景知识:/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx文件中的农林牧渔业总产值、农业、林业、牧业和渔业产值共计五张独立的sheet，列名都从第四行开始
列名有省份、2003年、2003年……一直到2022年等项,不存在汇总各个省份的全国项。产值单位为亿元，根据下列查询要求，输出实现代码
'''


def testss06():#计算出每个省份2022年相较2005年的农林牧渔业总产值增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='农林牧渔业总产值',engine='openpyxl',header=3)
     # 计算每个省份2022年相较2005年的农林牧渔业总产值增长率，并将结果存储在最后一列  
    df['增长率'] = df['2022年'] / df['2005年'] - 1  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取农林牧渔业总产值这张表  
    ws = book["农林牧渔业总产值"]  
  
    # 将结果写入原Excel文件的"农林牧渔业总产值"sheet中的最后一列  
    for col_num, value in enumerate(df['增长率'].values, start=1):  
        ws.cell(row=1, column=col_num+1, value=value)  # 写入增长率到最后一列  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def testss07():#计算出这20年中每年各个省份的牧业总产值平均值，结果存储在最后一行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='牧业',engine='openpyxl',header=3)
    # 计算这20年中每年各个省份的牧业总产值平均值  
    average_df = df.iloc[:, 1:].mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取农业这张表  
    ws = book["牧业"]  
  
    # 将结果写入原Excel文件的"牧业"sheet中的最后一行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=ws.max_row+1, column=col_num, value=value)  # 写入平均值到最后一行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def testss08():#计算出这20年中每年各个省份的渔业总产值平均值，结果存储在最后一行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='渔业',engine='openpyxl',header=3)
    # 计算每年各个省份的渔业总产值平均值  
    average_df = df.mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取渔业这张表  
    ws = book["渔业"]  
  
    # 将结果写入原Excel文件的"渔业"sheet中的最后一行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=ws.max_row+1, column=col_num, value=value)  # 写入平均值到最后一行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def testss09():#将各省份按照2018年的林业总产值降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
    # 将数据按照2018年的林业总产值降序排序  
    df = df.sort_values(by='2018年', ascending=False)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2018年'])  # 将排序后的数据写入"林业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def testss10():#将各省份按照2018年的林业总产值升序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
    # 将数据按照2018年的林业总产值升序排序  
    df = df.sort_values(by='2018年')  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2018年'])  # 将排序后的数据写入"林业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  




if __name__=='__main__':
    testss01()



