import pandas as pd
import numpy as np  
from openpyxl import load_workbook

'''
背景知识：/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx文件中
共有4个独立的sheet，分别为“民用汽车拥有量汇总”、“民用载货”、“民用载客”、“民用其他”。
“民用汽车拥有量汇总”、“民用载货”、“民用载客”的列名都从第5行开始，
列名有省份、2002年、2003年……一直到2020年等项，不存在汇总各个省份的全国项。
单位为万辆，根据下列查询要求，输出实现代码。
'''



def queryss01():#计算出每个省份这20年的民用汽车拥有量总和，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 计算每个省份这20年的民用汽车拥有量总和  
    df['总和'] = df.iloc[:, 1:].sum(axis=0)  # 修改了axis的值，从1改为0，以计算列的和  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的总和数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['总和'])  # 将总和写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss02():#计算出每个省份2020年相较于2002年的民用汽车拥有量增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 判断数据中是否存在'2002年'这一列  
    if '2002年' in df.columns:  
        # 计算每个省份2020年相较于2002年的民用汽车拥有量增长率  
        df['增长率'] = (df['2020年'] - df['2002年']) / df['2002年'] * 100  
    else:  
        print("数据中不存在'2002年'这一列")  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的增长率数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['增长率'])  # 将增长率写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss03():#计算出每个省份2012年相较于2002年的民用汽车拥有量增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 判断数据中是否存在'2002年'这一列  
    if '2002年' in df.columns:  
        # 计算每个省份2012年相较于2002年的民用汽车拥有量增长率  
        df['增长率'] = (df['2012年'] - df['2002年']) / df['2002年'] * 100  
    else:  
        print("数据中不存在'2002年'这一列")  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的增长率数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['增长率'])  # 将增长率写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss04():#计算出每个省份2020年相较于2012年的民用汽车拥有量增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 计算每个省份2020年相较于2012年的民用汽车拥有量增长率  
    df['增长率'] = (df['2020年'] - df['2012年']) / df['2012年'] * 100  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将计算后的增长率数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（在最后一列添加）  
    for index, row in df.iterrows():  
        ws.cell(row=index+1, column=len(df.columns)+1, value=row['增长率'])  # 将增长率写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss05():#将各省份按照2020年的民用汽车拥有量降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 按照2020年的民用汽车拥有量进行降序排序  
    df.sort_values('2020年', inplace=True, ascending=False)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将排序后的数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（覆盖原有数据）  
    for index, row in df.iterrows():  
        for col_num, value in enumerate(row.values, start=1):  
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)

def queryss06():#将各省份按照2020年的民用汽车拥有量升序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用汽车拥有量汇总',engine='openpyxl',header=4)
    # 对各省份按照2020年的民用汽车拥有量进行升序排序  
    df.sort_values('2020年', inplace=True)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用汽车拥有量汇总'这张表  
    ws = book["民用汽车拥有量汇总"]  
  
    # 将排序后的数据写入原Excel文件的'民用汽车拥有量汇总'sheet中（覆盖原有数据）  
    for index, row in df.iterrows():  
        for col_num, value in enumerate(row.values, start=1):  
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)

def queryss07():#计算出每个省份这20年的民用汽车拥有量总和，结果存储在最后一列，然后将各省所对应的行按照民用汽车拥有量总和降序排序。  
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'  
    df=pd.read_excel(file_name, sheet_name='民用载货',engine='openpyxl',header=4)  
     # 计算每个省份这20年的民用汽车拥有量总和，并将结果存储在最后一列    
    df['总和'] = df.iloc[:, 1:].sum(axis=1)    
    
    # 将各省所对应的行按照民用汽车拥有量总和降序排序    
    df.sort_values('总和', inplace=True, ascending=False)    
    
    # 加载原有的 Excel 文件    
    book = load_workbook(file_name)    
    
    # 获取'民用载货'这张表    
    ws = book["民用载货"]    
    
    # 将排序后的数据写入原Excel文件的'民用载货'sheet中（覆盖原有数据）    
    for index, row in df.iterrows():    
        for col_num, value in enumerate(row.values, start=1):    
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中    
    
    # 保存修改后的 Excel 文件    
    book.save(file_name)

def queryss08():#计算出每个省份这20年的民用汽车拥有量总和，结果存储在最后一列，然后将各省所对应的行按照民用汽车拥有量总和升序排序。
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用载货',engine='openpyxl',header=4)
     # 计算每个省份这20年的民用汽车拥有量总和，并将结果存储在最后一列  
    df['总和'] = df.iloc[:, 1:].sum(axis=1)  
  
    # 将各省所对应的行按照民用汽车拥有量总和升序排序  
    df.sort_values('总和', inplace=True)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用载货'这张表  
    ws = book["民用载货"]  
  
    # 将排序后的数据写入原Excel文件的'民用载货'sheet中（覆盖原有数据）  
    for index, row in df.iterrows():  
        for col_num, value in enumerate(row.values, start=1):  
            ws.cell(row=index+1, column=col_num, value=value)  # 将每一行的数据写入到相应的单元格中  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss09():#计算出这20年中每年各个省份的民用载货汽车拥有量的平均值，结果存储在最后一行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用载货',engine='openpyxl',header=4)
    # 计算每年各个省份的民用载货汽车拥有量的平均值  
    average_df = df.groupby(df.columns.tolist())[df.columns.tolist()].mean()  
  
    # 结果存储在最后一行  
    average_df.iloc[-1] = average_df.mean()  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用载货'这张表  
    ws = book["民用载货"]  
  
    # 将结果写入原Excel文件的'民用载货'sheet中的最后一行  
    for col_num, value in enumerate(average_df.iloc[-1].values, start=1):  
        ws.cell(row=average_df.shape[0]+1, column=col_num, value=value)  # 写入平均值到最后一行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)

def queryss10():#计算出这20年中每年各个省份的民用载货汽车拥有量的平均值，结果存储在第36行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/民用汽车拥有量汇总.xlsx'
    df=pd.read_excel(file_name, sheet_name='民用载货',engine='openpyxl',header=4)
    # 计算每年各个省份的民用载货汽车拥有量的平均值  
    average_df = df.iloc[:, 1:].mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取'民用载货'这张表  
    ws = book["民用载货"]  
  
    # 将结果写入原Excel文件的'民用载货'sheet中的第36行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=36, column=col_num, value=value)  # 写入平均值到第36行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

'''
背景知识:/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx文件中的农林牧渔业总产值、农业、林业、牧业和渔业产值共计五张独立的sheet，列名都从第四行开始
列名有省份、2003年、2003年……一直到2022年等项,不存在汇总各个省份的全国项。产值单位为亿元，根据下列查询要求，输出实现代码
'''

def queryss11():#计算出每个省份这20年的农林牧渔业总产值总和，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='农林牧渔业总产值',engine='openpyxl',header=3)
    # 计算每个省份这20年的农林牧渔业总产值总和，并将结果存储在最后一列  
    df['总和'] = df.iloc[:, 1:].sum(axis=1)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取农林牧渔业总产值这张表  
    ws = book["农林牧渔业总产值"]  
  
    # 将结果写入原Excel文件的"农林牧渔业总产值"sheet中的最后一列  
    for col_num, value in enumerate(df['总和'].values, start=1):  
        ws.cell(row=1, column=col_num+1, value=value)  # 写入总和到最后一列  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss12():#计算出每个省份2022年相较2003年的农林牧渔业总产值增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='农林牧渔业总产值',engine='openpyxl',header=3)
     # 计算每个省份2022年相较2003年的农林牧渔业总产值增长率，并将结果存储在最后一列  
    df['增长率'] = df['2022年'] / df['2003年'] - 1  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取农林牧渔业总产值这张表  
    ws = book["农林牧渔业总产值"]  
  
    # 将结果写入原Excel文件的"农林牧渔业总产值"sheet中的最后一列  
    for col_num, value in enumerate(df['增长率'].values, start=1):  
        ws.cell(row=1, column=col_num+1, value=value)  # 写入增长率到最后一列  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss13():#计算出这20年中每年各个省份的农业总产值平均值，结果存储在最后一行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='农业',engine='openpyxl',header=3)
    # 计算这20年中每年各个省份的农业总产值平均值  
    average_df = df.iloc[:, 1:].mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取农业这张表  
    ws = book["农业"]  
  
    # 将结果写入原Excel文件的"农业"sheet中的最后一行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=ws.max_row+1, column=col_num, value=value)  # 写入平均值到最后一行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss14():#计算出这20年中每年各个省份的林业总产值平均值，结果存储在最后一行
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
    # 计算每年各个省份的林业总产值平均值  
    average_df = df.mean(axis=0)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中的最后一行  
    for col_num, value in enumerate(average_df.values, start=1):  
        ws.cell(row=ws.max_row+1, column=col_num, value=value)  # 写入平均值到最后一行  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss15():#将各省份按照2022年的林业总产值降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
    # 将数据按照2022年的林业总产值降序排序  
    df = df.sort_values(by='2022年', ascending=False)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2022年'])  # 将排序后的数据写入"林业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss16():#将各省份按照2022年的林业总产值升序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
    # 将数据按照2022年的林业总产值升序排序  
    df = df.sort_values(by='2022年')  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2022年'])  # 将排序后的数据写入"林业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss17():#计算出每个省份这20年的林业总产值总和，结果存储在最后一列；然后将各省份按照这20年的林业总产值总和降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='林业',engine='openpyxl',header=3)
     # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取林业这张表  
    ws = book["林业"]  
  
    # 将结果写入原Excel文件的"林业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['总和'])  # 将排序后的数据写入"林业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name)  

def queryss18():#将各省份按照2022年的牧业总产值降序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='牧业',engine='openpyxl',header=3)
    # 将数据按照2022年的牧业总产值降序排序  
    df = df.sort_values(by='2022年', ascending=False)  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取牧业这张表  
    ws = book["牧业"]  
  
    # 将结果写入原Excel文件的"牧业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2022年'])  # 将排序后的数据写入"牧业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss19():#将各省份按照2022年的牧业总产值升序排序
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='牧业',engine='openpyxl',header=3)
     # 将数据按照2022年的牧业总产值升序排序  
    df = df.sort_values(by='2022年')  
  
    # 加载原有的 Excel 文件  
    book = load_workbook(file_name)  
  
    # 获取牧业这张表  
    ws = book.active  
  
    # 将结果写入原Excel文件的"牧业"sheet中，而不是创建新的Excel文件  
    ws.append(df.columns.values)  # 添加新的一列，包含列名  
    for i, row in df.iterrows():  
        ws.cell(row=i+1, column=1, value=row['2022年'])  # 将排序后的数据写入"牧业"sheet  
  
    # 保存修改后的 Excel 文件  
    book.save(file_name) 

def queryss20():#计算出每个省份2022年相较于2003年的牧业产值增长率，结果存储在最后一列
    file_name='/Users/yizhiyuan/academy/machine learning/Final Pre/农林牧渔业总产值.xlsx'
    df=pd.read_excel(file_name, sheet_name='牧业',engine='openpyxl',header=3)
    df['增长率'] = (df['2022年'] - df['2003年']) / df['2003年'] * 100  
    # 加载原有的 Excel 文件
    book = load_workbook(file_name)

    # 获取牧业这张表
    ws = book["牧业"]

    # 将增长率添加到牧业这张表的最后一列
    ws.cell(row=4, column=ws.max_column + 1, value="增长率")
    for i, value in enumerate(df["增长率"], start=5):
        ws.cell(row=i, column=ws.max_column, value=value)

    # 保存修改后的 Excel 文件
    book.save(file_name)
    


if __name__=='__main__':
    queryss02()



